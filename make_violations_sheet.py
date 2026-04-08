import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── 위반 데이터 ──────────────────────────────────────────────
violations = [
    # ── Page 1 : 오블리브 복합치료 ──────────────────────────────
    {
        "page": "오블리브 복합치료",
        "url": "obliv.kr/rehab/obliv-complex",
        "no": 1,
        "text": "일반 통증치료: 통증 개선되나 단기간 내 재발",
        "type": "타 치료법 비방·비교",
        "law": "의료법 제56조 제2항 제6호",
        "reason": "다른 치료법(일반 통증치료)을 부정적으로 묘사하며 비교표로 자원 우월성을 주장",
        "suggestion": "비교표 삭제. 자원 치료의 특징만 단독으로 기술",
    },
    {
        "page": "오블리브 복합치료",
        "url": "obliv.kr/rehab/obliv-complex",
        "no": 2,
        "text": "일반 통증치료: 주사·약물 의존도 높음",
        "type": "타 치료법 비방·비교",
        "law": "의료법 제56조 제2항 제6호",
        "reason": "일반 통증치료를 폄하하는 내용으로 타 치료법 비방",
        "suggestion": "삭제",
    },
    {
        "page": "오블리브 복합치료",
        "url": "obliv.kr/rehab/obliv-complex",
        "no": 3,
        "text": "일반 통증치료: 일시적 통증 완화에만 그침",
        "type": "타 치료법 비방·비교",
        "law": "의료법 제56조 제2항 제6호",
        "reason": "일반 치료법을 '일시적'으로 한정하며 비방",
        "suggestion": "삭제",
    },
    {
        "page": "오블리브 복합치료",
        "url": "obliv.kr/rehab/obliv-complex",
        "no": 4,
        "text": "3개월 이상 지속된 통증, 석회화·염증에 효과",
        "type": "치료 효과 단정",
        "law": "의료법 제56조 제2항 제2호",
        "reason": "특정 질환에 대한 치료 효과를 단정적으로 표현",
        "suggestion": "\"~에 도움이 될 수 있습니다\"로 변경",
    },
    {
        "page": "오블리브 복합치료",
        "url": "obliv.kr/rehab/obliv-complex",
        "no": 5,
        "text": "치료 시간은 짧게, 회복은 빠르게. 오랜 지속 효과",
        "type": "치료 효과 단정",
        "law": "의료법 제56조 제2항 제2호",
        "reason": "빠른 회복과 오랜 지속 효과를 보장하는 단정 표현",
        "suggestion": "\"치료 시간이 짧은 편이며, 지속적인 회복 관리 효과를 기대할 수 있습니다\"",
    },
    {
        "page": "오블리브 복합치료",
        "url": "obliv.kr/rehab/obliv-complex",
        "no": 6,
        "text": "재발률 감소 (복합치료 주요 기대 효과 항목)",
        "type": "치료 효과 단정",
        "law": "의료법 제56조 제2항 제2호",
        "reason": "재발률 감소 효과를 임상적 근거 없이 단정",
        "suggestion": "\"재발 예방에 도움이 될 수 있습니다\" 또는 근거 문헌 인용 후 명시",
    },
    {
        "page": "오블리브 복합치료",
        "url": "obliv.kr/rehab/obliv-complex",
        "no": 7,
        "text": "근본적인 통증을 개선하고, 자생회복력을 도우며 몸의 이완과 재생까지 유도하는 회복 메커니즘 완성",
        "type": "치료 효과 단정·과장",
        "law": "의료법 제56조 제2항 제2호",
        "reason": "\"완성\"이라는 확정적 표현으로 치료 효과를 보장",
        "suggestion": "\"~을 목표로 합니다\" 또는 \"~을 기대할 수 있습니다\"로 변경",
    },

    # ── Page 2 : ESWT 충격파치료 ────────────────────────────────
    {
        "page": "ESWT 충격파치료",
        "url": "obliv.kr/rehab/eswt",
        "no": 1,
        "text": "근본적인 통증을 완화하는 비침습 치료",
        "type": "치료 효과 단정",
        "law": "의료법 제56조 제2항 제2호",
        "reason": "\"근본적인\" 통증 완화를 단정적으로 기술",
        "suggestion": "\"통증 완화에 도움을 줄 수 있는 비침습 치료\"로 변경",
    },
    {
        "page": "ESWT 충격파치료",
        "url": "obliv.kr/rehab/eswt",
        "no": 2,
        "text": "안정적, 효과적이며 부작용 위험이 매우 낮은 치료",
        "type": "효과 과장·부작용 정보 축소",
        "law": "의료법 제56조 제2항 제2호",
        "reason": "\"효과적\"이라는 단정 표현 + \"부작용 매우 낮음\" 단정으로 부작용 정보 축소",
        "suggestion": "참고 문헌 출처를 명시하고 \"연구에 따르면 ~한 것으로 보고되어 있습니다\"로 변경",
    },
    {
        "page": "ESWT 충격파치료",
        "url": "obliv.kr/rehab/eswt",
        "no": 3,
        "text": "소염제, 국소 주사치료, 재활치료 등으로도 호전되지 않던 관절·힘줄·인대 등의 손상을 회복시키는 기술입니다",
        "type": "타 치료법 비교·효과 단정",
        "law": "의료법 제56조 제2항 제6호",
        "reason": "타 치료법(소염제·주사·재활치료)이 효과 없는 경우도 치료된다고 단정하며 우월성 주장",
        "suggestion": "삭제 또는 \"기존 치료에 반응이 적은 경우 대안으로 고려할 수 있습니다\" 수준으로 완화",
    },
    {
        "page": "ESWT 충격파치료",
        "url": "obliv.kr/rehab/eswt",
        "no": 4,
        "text": "대부분의 환자분들은 치료 후 통증이 감소하거나 사라지는 것을 경험합니다",
        "type": "치료 효과 단정",
        "law": "의료법 제56조 제2항 제2호",
        "reason": "\"대부분\"이라는 표현으로 치료 효과를 사실상 보장",
        "suggestion": "삭제하거나 인용 논문 명시 후 \"~으로 보고된 바 있습니다\"로 변경",
    },
    {
        "page": "ESWT 충격파치료",
        "url": "obliv.kr/rehab/eswt",
        "no": 5,
        "text": "시술 후 바로 일상생활이 가능하며 별도의 회복 기간이 거의 없어 반복 치료를 받으실 수 있습니다",
        "type": "치료 효과 단정",
        "law": "의료법 제56조 제2항 제2호",
        "reason": "시술 직후 즉시 일상 복귀를 단정적으로 보장",
        "suggestion": "\"대부분의 경우 일상생활에 큰 지장이 없으나, 개인별 차이가 있을 수 있습니다\"",
    },
    {
        "page": "ESWT 충격파치료",
        "url": "obliv.kr/rehab/eswt",
        "no": 6,
        "text": "일정 기간 반복 유지치료 시 재발률 감소 (치료 주기 가이드 항목)",
        "type": "치료 효과 단정",
        "law": "의료법 제56조 제2항 제2호",
        "reason": "재발률 감소를 임상적 근거 없이 단정",
        "suggestion": "\"재발 예방에 도움이 될 수 있습니다\" + 관련 근거 인용",
    },

    # ── Page 3 : 산후 도수치료 ───────────────────────────────────
    {
        "page": "산후 도수치료",
        "url": "obliv.kr/rehab/postpartum-therapy",
        "no": 1,
        "text": "평생 건강을 결정하는 산후 6개월의 골든타임",
        "type": "과장 광고·공포심 자극",
        "law": "의료법 제56조 제2항 제2호",
        "reason": "\"평생 건강을 결정\"한다는 과장 표현으로 환자 불안감 조성",
        "suggestion": "\"산후 6개월은 신체 회복에 중요한 시기입니다\"로 순화",
    },
    {
        "page": "산후 도수치료",
        "url": "obliv.kr/rehab/postpartum-therapy",
        "no": 2,
        "text": "골든타임 내 치료는 산후풍과 만성 통증을 예방하는 가장 확실한 방법입니다",
        "type": "최상급 표현·효과 보장",
        "law": "의료법 제56조 제2항 제2호",
        "reason": "\"가장 확실한 방법\"이라는 최상급 표현으로 예방 효과를 보장",
        "suggestion": "\"도움이 될 수 있는 방법 중 하나입니다\"로 변경",
    },
    {
        "page": "산후 도수치료",
        "url": "obliv.kr/rehab/postpartum-therapy",
        "no": 3,
        "text": "골든타임을 놓쳐 인대가 다시 단단해지면 불균형한 정렬이 그대로 고착될 수 있습니다",
        "type": "공포심 자극",
        "law": "의료광고 심의기준 (과장·공포 유발)",
        "reason": "치료하지 않으면 문제가 고착된다며 공포심 조장",
        "suggestion": "\"치료 시기가 이를수록 회복에 유리할 수 있습니다\" 수준으로 순화",
    },
    {
        "page": "산후 도수치료",
        "url": "obliv.kr/rehab/postpartum-therapy",
        "no": 4,
        "text": "오직 회복에만 전념할 수 있는 최적의 환경을 선사합니다",
        "type": "최상급 표현",
        "law": "의료광고 심의기준 (최상급 표현 금지)",
        "reason": "\"최적의 환경\"이라는 최상급 표현",
        "suggestion": "\"회복에 집중할 수 있는 독립적인 치료 환경을 갖추고 있습니다\"로 변경",
    },

    # ── Page 4 : 문제성발톱 – 발톱무좀 ─────────────────────────────
    {
        "page": "발톱무좀",
        "url": "obliv.kr/foot?category=infection",
        "no": 1,
        "text": "레이저 광선이 병변 깊숙이 도달해 약물이 닿기 어려운 부위까지 무좀균을 제거하고 정상 세포의 재생을 촉진합니다",
        "type": "치료 효과 단정",
        "law": "의료법 제56조 제2항 제2호",
        "reason": "균 제거 및 세포 재생 효과를 단정적으로 표현",
        "suggestion": "\"~에 도달해 무좀균 억제 및 세포 재생 촉진을 기대할 수 있습니다\"로 변경",
    },
    {
        "page": "발톱무좀",
        "url": "obliv.kr/foot?category=infection",
        "no": 2,
        "text": "통증과 부작용 거의 없어 시술 직후 일상생활 가능",
        "type": "부작용 정보 축소·효과 단정",
        "law": "의료법 제56조 제2항 제2호",
        "reason": "부작용이 \"거의 없다\"는 표현으로 부작용 정보 축소, 시술 직후 일상생활 가능을 단정",
        "suggestion": "\"개인에 따라 경미한 불편감이 있을 수 있으나, 대부분 일상생활 가능합니다\"로 변경",
    },
    {
        "page": "발톱무좀",
        "url": "obliv.kr/foot?category=infection",
        "no": 3,
        "text": "고혈압·당뇨 등 기저질환자도 동시 치료 가능",
        "type": "특수 대상 치료 가능 단정",
        "law": "의료법 제56조 제2항 제2호",
        "reason": "기저질환자에 대한 치료 안전성을 조건 없이 단정",
        "suggestion": "\"의료진 판단에 따라 가능할 수 있으며, 사전 진료를 권장드립니다\"로 변경",
    },
    {
        "page": "발톱무좀",
        "url": "obliv.kr/foot?category=infection",
        "no": 4,
        "text": "임신부·수유부도 치료 가능",
        "type": "특수 대상 치료 가능 단정",
        "law": "의료법 제56조 제2항 제2호",
        "reason": "임신부·수유부에 대한 안전성을 무조건적으로 단정",
        "suggestion": "\"의료진 진단 후 개인 상태에 따라 가능할 수 있습니다\"로 변경",
    },
    {
        "page": "발톱무좀",
        "url": "obliv.kr/foot?category=infection",
        "no": 5,
        "text": "곰팡이균만 선택적으로 사멸시키는 치료입니다",
        "type": "치료 효과 단정",
        "law": "의료법 제56조 제2항 제2호",
        "reason": "선택적 사멸이라는 효과를 단정적으로 표현",
        "suggestion": "\"곰팡이균을 선택적으로 타겟팅하는 방식으로 작용합니다\" (기기 원리 설명 수준으로 유지)",
    },
    {
        "page": "발톱무좀",
        "url": "obliv.kr/foot?category=infection",
        "no": 6,
        "text": "방치 시 감염 부위가 빠르게 확산되고 치료 기간이 길어질 수 있습니다",
        "type": "공포심 자극",
        "law": "의료광고 심의기준 (과장·공포 유발)",
        "reason": "방치 시 빠른 확산을 강조하여 치료를 강요하는 불안감 유발",
        "suggestion": "\"조기에 진료를 받으시면 치료가 수월할 수 있습니다\"로 순화",
    },

    # ── Page 5 : 문제성발톱 – 내성발톱 ─────────────────────────────
    {
        "page": "내성발톱",
        "url": "obliv.kr/foot?category=ingrown",
        "no": 1,
        "text": "걸을 때마다 찌르고, 부어오르고, 진물이 나는 상태라면 이미 치료가 늦은 단계일 수 있습니다",
        "type": "공포심 자극",
        "law": "의료광고 심의기준 (과장·공포 유발)",
        "reason": "\"이미 치료가 늦은 단계\"라는 표현으로 환자 공포심 자극",
        "suggestion": "\"진료가 필요한 상태일 수 있습니다\"로 순화",
    },
    {
        "page": "내성발톱",
        "url": "obliv.kr/foot?category=ingrown",
        "no": 2,
        "text": "방치할 경우 염증 확산 및 세균 감염으로 발전해 치료가 보다 복잡해질 수 있습니다",
        "type": "공포심 자극",
        "law": "의료광고 심의기준 (과장·공포 유발)",
        "reason": "방치 시 세균 감염으로 발전한다고 강조하여 불안감 유발",
        "suggestion": "\"증상이 지속되면 전문의 진료를 받으시는 것을 권장드립니다\"로 순화",
    },
    {
        "page": "내성발톱",
        "url": "obliv.kr/foot?category=ingrown",
        "no": 3,
        "text": "국내 다수 정형외과에서 사용중인 검증 의료기기",
        "type": "근거 없는 우수성 주장",
        "law": "의료광고 심의기준 (허위·과장)",
        "reason": "\"검증\"이라는 표현과 \"국내 다수 정형외과\" 사용 주장의 구체적 근거 미제시",
        "suggestion": "\"식약처 인증 의료기기\"로 대체하거나 구체적 근거 자료 인용 필요",
    },
    {
        "page": "내성발톱",
        "url": "obliv.kr/foot?category=ingrown",
        "no": 4,
        "text": "곡률이 100도 이하라면, 통증이 없어도 의학적 치료가 필요한 상태입니다",
        "type": "치료 필요성 단정",
        "law": "의료법 제56조 제2항 제2호",
        "reason": "증상 없는 경우에도 치료가 반드시 필요하다고 단정하여 불필요한 의료 이용 유도 가능",
        "suggestion": "\"의료진 진단 후 치료 여부를 결정하시기 바랍니다\"로 변경",
    },
]

# ── 스타일 설정 ────────────────────────────────────────────────
PAGE_COLORS = {
    "오블리브 복합치료": "D6E4F0",
    "ESWT 충격파치료":   "D5F0D6",
    "산후 도수치료":     "FCE4D6",
    "발톱무좀":          "EDE0F5",
    "내성발톱":          "FFF0CC",
}
TYPE_COLORS = {
    "타 치료법 비방·비교":      "FF8080",
    "타 치료법 비교·효과 단정": "FF9999",
    "치료 효과 단정":           "FFB347",
    "치료 효과 단정·과장":      "FFB347",
    "효과 과장·부작용 정보 축소": "FFC080",
    "부작용 정보 축소·효과 단정": "FFC080",
    "최상급 표현":              "FFE066",
    "최상급 표현·효과 보장":    "FFE066",
    "과장 광고·공포심 자극":    "FF9966",
    "공포심 자극":              "FF9966",
    "특수 대상 치료 가능 단정": "98D8C8",
    "근거 없는 우수성 주장":    "C0B0FF",
    "치료 필요성 단정":         "FFB347",
}

HEADER_FILL   = "1E3A5F"
HEADER_FONT   = "FFFFFF"
THIN          = Side(style="thin", color="CCCCCC")
BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def cell_style(ws, row, col, value, bold=False, fill_hex=None, wrap=True, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="맑은 고딕", bold=bold,
                  color="FFFFFF" if fill_hex == HEADER_FILL else "000000", size=9)
    if fill_hex:
        c.fill = PatternFill("solid", fgColor=fill_hex)
    c.alignment = Alignment(wrap_text=wrap, vertical="top",
                             horizontal=align if align != "left" else "left")
    c.border = BORDER
    return c

COLS = ["페이지", "URL", "No.", "위반 표현 (원문)", "위반 유형", "관련 법령", "위반 사유", "수정 제안"]
WIDTHS = [18, 34, 5, 52, 24, 30, 44, 44]

def build_sheet(ws, rows, title):
    ws.title = title
    ws.freeze_panes = "A2"

    # 제목행
    for col_i, (col_name, width) in enumerate(zip(COLS, WIDTHS), 1):
        cell_style(ws, 1, col_i, col_name, bold=True, fill_hex=HEADER_FILL)
        ws.column_dimensions[get_column_letter(col_i)].width = width

    for r_i, row in enumerate(rows, 2):
        page_color = PAGE_COLORS.get(row["page"], "FFFFFF")
        type_color = TYPE_COLORS.get(row["type"], "FFFFFF")
        values = [
            row["page"], row["url"], row["no"],
            row["text"], row["type"], row["law"],
            row["reason"], row["suggestion"],
        ]
        for col_i, val in enumerate(values, 1):
            fill = (page_color if col_i <= 3
                    else type_color if col_i == 5
                    else None)
            bold = col_i == 3
            cell_style(ws, r_i, col_i, val, bold=bold, fill_hex=fill)

    ws.row_dimensions[1].height = 20
    for r_i in range(2, len(rows) + 2):
        ws.row_dimensions[r_i].height = 58


# ── 전체 요약 시트 ─────────────────────────────────────────────
ws_all = wb.active
build_sheet(ws_all, violations, "전체 요약")

# ── 페이지별 시트 ──────────────────────────────────────────────
pages_order = ["오블리브 복합치료", "ESWT 충격파치료", "산후 도수치료", "발톱무좀", "내성발톱"]
for page_name in pages_order:
    rows = [v for v in violations if v["page"] == page_name]
    ws = wb.create_sheet(title=page_name)
    build_sheet(ws, rows, page_name)

OUT = "/home/user/Marketing/의료법_위반표현_오블리브.xlsx"
wb.save(OUT)
print(f"저장 완료: {OUT}")
print(f"총 위반 항목: {len(violations)}건")
for p in pages_order:
    cnt = sum(1 for v in violations if v["page"] == p)
    print(f"  {p}: {cnt}건")
